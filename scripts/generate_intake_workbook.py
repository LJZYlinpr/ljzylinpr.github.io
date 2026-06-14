from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "files" / "主页信息采集模板.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
SECTION_FILL = PatternFill("solid", fgColor="D9F2EF")
NOTE_FILL = PatternFill("solid", fgColor="FFF6E8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row_num):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BOX


def style_table(ws, start_row, end_row, widths):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_notes(ws, title, notes):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = notes
    ws["A2"].fill = NOTE_FILL
    ws["A2"].border = BOX
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 78


def make_directory(ws):
    ws.title = "目录"
    ws["A1"] = "主页资料填写目录"
    ws["A1"].font = Font(bold=True, size=16)
    headers = ["工作表", "用途", "你需要填写什么", "是否必须", "备注"]
    for col, header in zip("ABCDE", headers):
        ws[f"{col}3"] = header
    style_header(ws, 3)
    rows = [
        ("目录", "总说明", "查看填写顺序和各页作用", "是", ""),
        ("个人信息", "侧边栏和主页首屏", "姓名、邮箱、头像、简介、主页主标题", "是", ""),
        ("首页内容", "主页模块", "首页分区文案、兴趣标签、统计卡片", "是", ""),
        ("新闻动态", "主页时间线", "每条新闻的时间、标题、说明、链接", "建议", ""),
        ("论文信息", "论文页面和主页精选论文", "论文标题、作者、简介、配图、链接", "是", "每篇论文建议配一张图"),
        ("团队信息", "合作成员或团队展示", "姓名、身份、简介、照片、链接", "建议", "每人或每组一张图"),
        ("助教经历", "Teaching 页面", "学期、课程、角色、简介", "建议", "不需要配图"),
        ("图片素材", "生活/杂项图墙", "图片标题、路径、时间、人物、地点", "可选", ""),
        ("访客地图", "主页访客地图模块", "地图标题、展示方式、是否展示国家/城市统计", "建议", "我后续会按更稳妥的隐私方案实现"),
        ("素材清单", "总核对", "确认每类图片和文件是否准备好", "建议", ""),
    ]
    for idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value).border = BOX
    style_table(ws, 4, 13, {"A": 16, "B": 24, "C": 42, "D": 12, "E": 34})
    ws.freeze_panes = "A4"


def make_profile(ws):
    ws.title = "个人信息"
    add_notes(ws, "个人信息", "这一页用于填写你的基本资料、头像、联系方式、主页首屏标题。图片路径请直接填你本地图片的绝对路径，或者把图片放进我给你建好的素材文件夹后填写对应路径。")
    start = 4
    headers = ["字段代号", "显示名称", "填写内容", "是否必填", "示例"]
    for col, header in zip("ABCDE", headers):
        ws[f"{col}{start}"] = header
    style_header(ws, start)
    rows = [
        ("name_en", "英文名", "Peiran Lin", "是", "Peiran Lin"),
        ("name_zh", "中文名", "", "建议", "蔺培然"),
        ("main_title", "主页主身份", "", "是", "西湖大学生命科学博士生"),
        ("sub_title", "副标题", "", "建议", "AI for Biology / Spatial Omics"),
        ("institution", "机构", "Westlake University", "是", "Westlake University"),
        ("location", "所在地", "Hangzhou, China", "是", "Hangzhou, China"),
        ("email", "邮箱", "", "是", "linpeiran@westlake.edu.cn"),
        ("bio_short", "侧边栏短简介", "", "是", "一句话简介"),
        ("hero_title", "首页首屏大标题", "", "是", "一句有辨识度的主标题"),
        ("hero_intro", "首页首屏简介", "", "是", "2到4句"),
        ("avatar_path", "头像图片路径", "", "是", "/home/linpeiran/homepage/linpeiran_homepage.github.io/files/网站素材收集/01_个人头像/头像.jpg"),
        ("scholar_url", "Google Scholar", "", "建议", "https://scholar.google.com/..."),
        ("github_url", "GitHub", "", "建议", "https://github.com/LJZYlinpr"),
        ("researchgate_url", "ResearchGate", "", "可选", "https://www.researchgate.net/..."),
        ("orcid_url", "ORCID", "", "可选", "https://orcid.org/..."),
        ("linkedin_url", "LinkedIn", "", "可选", "https://www.linkedin.com/..."),
        ("primary_lab", "主要实验室", "", "建议", "Zeng Lab"),
        ("primary_lab_url", "主要实验室链接", "", "建议", "https://..."),
    ]
    for idx, row in enumerate(rows, start=start + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    style_table(ws, start + 1, start + len(rows), {"A": 18, "B": 22, "C": 58, "D": 12, "E": 38})
    ws.freeze_panes = "A5"


def make_homepage(ws):
    ws.title = "首页内容"
    add_notes(ws, "首页内容", "这里控制主页各个模块的文字。每一行填一个内容槽位，后续我会按这些内容排版到首页。")
    start = 4
    headers = ["模块", "位置代号", "填写内容", "图片路径", "排序", "备注"]
    for col, header in zip("ABCDEF", headers):
        ws[f"{col}{start}"] = header
    style_header(ws, start)
    rows = [
        ("首屏", "眉题", "", "", 1, "大标题上面的小字"),
        ("首屏", "大标题", "", "", 2, "首页最醒目的标题"),
        ("首屏", "简介", "", "", 3, "首屏段落"),
        ("快照卡片", "卡片1", "", "", 1, "例如：项目 | PhD in Life Sciences"),
        ("快照卡片", "卡片2", "", "", 2, ""),
        ("快照卡片", "卡片3", "", "", 3, ""),
        ("快照卡片", "卡片4", "", "", 4, ""),
        ("统计卡片", "统计1", "", "", 1, "例如：4 | 已展示论文"),
        ("统计卡片", "统计2", "", "", 2, ""),
        ("统计卡片", "统计3", "", "", 3, ""),
        ("统计卡片", "统计4", "", "", 4, ""),
        ("关于我", "段落1", "", "", 1, ""),
        ("关于我", "段落2", "", "", 2, ""),
        ("研究兴趣", "标签集合", "", "", 1, "多个标签用 | 分隔"),
        ("团队", "首页团队引导语", "", "", 1, "团队卡片上方的一句话"),
    ]
    for idx, row in enumerate(rows, start=start + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    style_table(ws, start + 1, start + len(rows), {"A": 16, "B": 18, "C": 56, "D": 32, "E": 10, "F": 36})
    ws.freeze_panes = "A5"


def make_news(ws):
    ws.title = "新闻动态"
    add_notes(ws, "新闻动态", "每行一条新闻，时间尽量写准确日期。主页只会展示较新的几条。")
    start = 4
    headers = ["日期", "标题", "说明", "链接", "是否展示到首页", "优先级"]
    for col, header in zip("ABCDEF", headers):
        ws[f"{col}{start}"] = header
    style_header(ws, start)
    style_table(ws, start + 1, start + 20, {"A": 16, "B": 34, "C": 58, "D": 28, "E": 16, "F": 10})
    ws.freeze_panes = "A5"


def make_publications(ws):
    ws.title = "论文信息"
    add_notes(ws, "论文信息", "每篇论文一行。请务必填写简介和配图路径，这是你特别要求保留的基础展示模式。")
    start = 4
    headers = [
        "论文编号", "年份", "状态", "标题", "作者", "期刊/会议", "DOI或链接", "你的身份",
        "一句话简介", "详细简介（80-150字）", "配图路径", "配图说明", "裁剪备注", "是否首页展示", "排序"
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    style_table(
        ws,
        start + 1,
        start + 25,
        {
            "A": 12, "B": 10, "C": 12, "D": 40, "E": 36, "F": 22, "G": 24,
            "H": 14, "I": 28, "J": 52, "K": 34, "L": 24, "M": 18, "N": 14, "O": 10
        },
    )
    dv = DataValidation(type="list", formula1='"已发表,预印本,投稿中,准备中,会议论文"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{start+1}:C{start+100}")
    ws.freeze_panes = "A5"


def make_team(ws):
    ws.title = "团队信息"
    add_notes(ws, "团队信息", "每行一个人或一个小组。这里默认每个条目都有一张简洁配图和一句简介。")
    start = 4
    headers = ["编号", "姓名", "身份/角色", "单位", "关系类型", "一句话简介", "照片路径", "图片说明", "链接", "是否首页展示", "排序"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    style_table(ws, start + 1, start + 30, {"A": 10, "B": 16, "C": 18, "D": 20, "E": 14, "F": 40, "G": 32, "H": 22, "I": 24, "J": 14, "K": 10})
    ws.freeze_panes = "A5"


def make_teaching(ws):
    ws.title = "助教经历"
    add_notes(ws, "助教经历", "这一页不需要图片。每行写一段助教/课程/朋辈辅导经历即可。")
    start = 4
    headers = ["学期", "年份", "课程或活动", "角色", "机构", "一句话简介", "补充说明", "排序"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    style_table(ws, start + 1, start + 25, {"A": 14, "B": 10, "C": 34, "D": 16, "E": 22, "F": 34, "G": 42, "H": 10})
    ws.freeze_panes = "A5"


def make_gallery(ws):
    ws.title = "图片素材"
    add_notes(ws, "图片素材", "这个工作表用于生活照片墙、杂项页面、活动图集。每行一张图。")
    start = 4
    headers = ["分类", "图片标题", "图片路径", "时间", "人物", "地点", "是否展示", "排序"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    style_table(ws, start + 1, start + 40, {"A": 18, "B": 28, "C": 36, "D": 16, "E": 30, "F": 28, "G": 12, "H": 10})
    ws.freeze_panes = "A5"


def make_visitormap(ws):
    ws.title = "访客地图"
    add_notes(ws, "访客地图", "这里是你想要的主页访客地图模块需求表。我后续会优先用更稳妥的隐私友好方式做，不默认长期保存原始IP。")
    start = 4
    headers = ["设置项", "填写内容", "是否必填", "推荐值", "备注"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    rows = [
        ("是否启用地图模块", "是", "是", "是", ""),
        ("地图标题", "", "是", "Visitors Around The World", ""),
        ("地图副标题", "", "建议", "Approximate locations from recent visits", ""),
        ("你偏好的方案", "", "是", "第三方统计 / 自建轻量接口 / Cloudflare Worker", "三选一或写偏好"),
        ("是否长期保存原始IP", "否", "是", "否", "强烈建议不要长期保存原始IP"),
        ("展示粒度", "", "是", "国家 或 城市", "国家更稳妥"),
        ("数据保留天数", "", "建议", "30", ""),
        ("是否展示最近访问人数", "", "建议", "是", ""),
        ("是否展示国家排行", "", "建议", "是", ""),
        ("是否展示城市排行", "", "可选", "否", ""),
        ("是否展示实时闪烁点", "", "可选", "否", "静态聚合更稳妥"),
        ("合规提示文案", "", "可选", "This site uses approximate visit geolocation for aggregate display.", ""),
        ("你已有的服务", "", "可选", "", "写服务名，不要写密钥"),
    ]
    for idx, row in enumerate(rows, start=start + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    style_table(ws, start + 1, start + len(rows), {"A": 24, "B": 36, "C": 12, "D": 26, "E": 44})
    ws.freeze_panes = "A5"


def make_assets(ws):
    ws.title = "素材清单"
    add_notes(ws, "素材清单", "这里用来核对哪些图和文件已经准备好。你也可以把这里当成最终交付前的检查单。")
    start = 4
    headers = ["素材类型", "素材名称", "用于哪里", "文件路径或链接", "状态", "备注"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(start, idx, header)
    style_header(ws, start)
    rows = [
        ("图片", "个人头像", "侧边栏 + 首页", "", "待准备", ""),
        ("图片", "论文配图1", "论文页/首页", "", "待准备", ""),
        ("图片", "团队照片或成员照片", "团队页/首页", "", "待准备", ""),
        ("文件", "CV PDF", "个人资料", "", "可选", ""),
        ("图片", "生活图集", "Miscellaneous", "", "可选", ""),
        ("服务", "访客地图方案", "主页地图模块", "", "待决定", ""),
    ]
    for idx, row in enumerate(rows, start=start + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    style_table(ws, start + 1, start + 20, {"A": 14, "B": 24, "C": 24, "D": 42, "E": 12, "F": 34})
    dv = DataValidation(type="list", formula1='"待准备,已准备,可选,跳过,待决定"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{start+1}:E{start+100}")
    ws.freeze_panes = "A5"


def main():
    wb = Workbook()
    make_directory(wb.active)
    make_profile(wb.create_sheet())
    make_homepage(wb.create_sheet())
    make_news(wb.create_sheet())
    make_publications(wb.create_sheet())
    make_team(wb.create_sheet())
    make_teaching(wb.create_sheet())
    make_gallery(wb.create_sheet())
    make_visitormap(wb.create_sheet())
    make_assets(wb.create_sheet())
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
